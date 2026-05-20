import os
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def export_to_csv(df, filepath):
    """
    Saves a DataFrame as a CSV file.
    """
    # Ensure export directory exists
    export_dir = os.path.dirname(filepath)
    if export_dir and not os.path.exists(export_dir):
        os.makedirs(export_dir, exist_ok=True)
        
    df.to_csv(filepath, index=False, encoding="utf-8-sig")
    return filepath

def export_to_excel(df, filepath):
    """
    Saves a DataFrame as an Excel file with premium custom styling.
    """
    # Ensure export directory exists
    export_dir = os.path.dirname(filepath)
    if export_dir and not os.path.exists(export_dir):
        os.makedirs(export_dir, exist_ok=True)
        
    # Write to Excel using Pandas + OpenPyXL
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="PDF Markups")
        
        # Get sheet to apply styles
        workbook = writer.book
        worksheet = writer.sheets["PDF Markups"]
        
        # Styles definition
        # Header style: Navy Blue background, white bold text
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data row style
        data_font = Font(name="Segoe UI", size=10)
        data_align_left = Alignment(horizontal="left", vertical="center")
        data_align_center = Alignment(horizontal="center", vertical="center")
        
        # Borders
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        # Color codes for Status column
        # Status options: Pending, In Progress, Resolved, Closed, Rejected
        status_fills = {
            "Pending": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),     # Soft amber
            "In Progress": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"), # Soft blue
            "Resolved": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),    # Soft emerald
            "Closed": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),      # Soft gray
            "Rejected": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),    # Soft red
        }
        status_fonts = {
            "Pending": Font(name="Segoe UI", size=10, color="B45309", bold=True),
            "In Progress": Font(name="Segoe UI", size=10, color="1D4ED8", bold=True),
            "Resolved": Font(name="Segoe UI", size=10, color="047857", bold=True),
            "Closed": Font(name="Segoe UI", size=10, color="374151", bold=True),
            "Rejected": Font(name="Segoe UI", size=10, color="B91C1C", bold=True),
        }
        
        # Color codes for Priority column
        priority_fills = {
            "Critical": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            "High": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
            "Medium": PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"),
            "Low": PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid"),
        }
        priority_fonts = {
            "Critical": Font(name="Segoe UI", size=10, color="991B1B", bold=True),
            "High": Font(name="Segoe UI", size=10, color="9A3412", bold=True),
            "Medium": Font(name="Segoe UI", size=10, color="1E40AF"),
            "Low": Font(name="Segoe UI", size=10, color="166534"),
        }
        
        # Format Headers
        worksheet.row_dimensions[1].height = 28
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            
        # Identify columns by name to apply specific alignment/coloring
        col_names = list(df.columns)
        status_col_idx = col_names.index("Status") + 1 if "Status" in col_names else -1
        priority_col_idx = col_names.index("Priority") + 1 if "Priority" in col_names else -1
        page_col_idx = col_names.index("Page Number") + 1 if "Page Number" in col_names else -1
        type_col_idx = col_names.index("Annotation Type") + 1 if "Annotation Type" in col_names else -1
        
        # Apply body styles
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 20
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                
                # Alignments
                if col_num in (page_col_idx, type_col_idx, status_col_idx, priority_col_idx):
                    cell.alignment = data_align_center
                else:
                    cell.alignment = data_align_left
                    
                # Status column styling
                if col_num == status_col_idx:
                    val = str(cell.value)
                    if val in status_fills:
                        cell.fill = status_fills[val]
                        cell.font = status_fonts[val]
                        
                # Priority column styling
                if col_num == priority_col_idx:
                    val = str(cell.value)
                    if val in priority_fills:
                        cell.fill = priority_fills[val]
                        cell.font = priority_fonts[val]
                        
        # Auto-adjust column widths with safety margin
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # Find the max string length in the column
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            
            # Apply padding, min-width, max-width limit to prevent overly wide columns
            col_width = min(max(max_len + 4, 12), 45)
            worksheet.column_dimensions[col_letter].width = col_width
            
    return filepath
